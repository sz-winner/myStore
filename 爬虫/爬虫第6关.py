import openpyxl
wb = openpyxl.Workbook()  # 创建一个excel文件
sheet = wb.active  # 激活excel文件的工作表
sheet.title = 'new title'  # 重命名工作表，可略过
sheet['A1'] = '漫威宇宙'  # 在A1单元格写入内容
rows = [['美国队长', '钢铁侠', '蜘蛛侠'], ['是', '漫威', '宇宙', '经典', '人物']]
for i in rows:
    sheet.append(i)  # 逐行写入内容
print(rows)
wb.save('D:\\Python\\test\\Marvel.xlsx')  # 保存文件

wb = openpyxl.load_workbook('D:\\Python\\test\\Marvel.xlsx')  # 读取excel文件
sheet = wb['new title']  # 定位工作表
sheetname = wb.sheetnames  # 获取工作表名称
print(sheetname)  # 打印工作表名称
A1_cell = sheet['A1']  # 定位单元格位置
A1_value = A1_cell.value  #获取单元格的值
print(A1_value)

'''
练习一：以下代码将爬到的周杰伦歌曲信息写到excel表格
'''
import requests, openpyxl
# 创建工作薄
wb = openpyxl.Workbook()
# 获取工作薄的活动表
sheet = wb.active
# 工作表重命名
sheet.title = 'lyrics'

sheet['A1'] = '歌曲名'  # 加表头，给A1单元格赋值
sheet['B1'] = '所属专辑'  # 加表头，给B1单元格赋值
sheet['C1'] = '播放时长'  # 加表头，给C1单元格赋值
sheet['D1'] = '播放链接'  # 加表头，给D1单元格赋值

url = 'https://c.y.qq.com/soso/fcgi-bin/client_search_cp'
for x in range(5):
    params = {
        'ct': '24',
        'qqmusic_ver': '1298',
        'new_json': '1',
        'remoteplace': 'txt.yqq.song',
        'searchid': '64405487069162918',
        't': '0',
        'aggr': '1',
        'cr': '1',
        'catZhida': '1',
        'lossless': '0',
        'flag_qc': '0',
        'p': str(x + 1),
        'n': '20',
        'w': '周杰伦',
        'g_tk': '5381',
        'loginUin': '0',
        'hostUin': '0',
        'format': 'json',
        'inCharset': 'utf8',
        'outCharset': 'utf-8',
        'notice': '0',
        'platform': 'yqq.json',
        'needNewCode': '0'
    }

    res_music = requests.get(url, params=params)
    json_music = res_music.json()
    list_music = json_music['data']['song']['list']
    for music in list_music:
        # 以name为键，查找歌曲名，把歌曲名赋值给name
        name = music['name']
        # 查找专辑名，把专辑名赋给album
        album = music['album']['name']
        # 查找播放时长，把时长赋值给time
        time = music['interval']
        # 查找播放链接，把链接赋值给link
        link = 'https://y.qq.com/n/yqq/song/' + str(music['mid']) + '.html\n\n'
        # 把name、album、time和link写成列表，用append函数多行写入Excel
        sheet.append([name, album, time, link])
        print('歌曲名：' + name + '\n' + '所属专辑:' + album + '\n' + '播放时长:' +
              str(time) + '\n' + '播放链接:' + link)

# 最后保存并命名这个Excel文件
wb.save('Jay.xlsx')

'''
练习二：爬取豆瓣TOP250的电影名/评分/推荐语/链接，并分别用csv和excel格式把它们存储下来
'''